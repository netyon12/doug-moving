"""
Módulo de Fretados - VERSÃO FINAL CORRETA
==========================================

Gerenciamento e visualização de fretados.
CORREÇÃO FINAL: Baseado na estrutura REAL do banco de dados.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, jsonify
from flask_login import login_required, current_user
from datetime import datetime, date
from sqlalchemy import func, or_, and_
from io import StringIO, BytesIO
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .. import db
from ..models import (
    Fretado, Empresa, Planta, Bloco,
    Gerente, Supervisor
)
from ..decorators import permission_required

from .admin import admin_bp


@admin_bp.route('/fretados')
@login_required
@permission_required(['admin', 'gerente', 'supervisor', 'operador'])
def fretados():
    """
    Lista todos os fretados (colaboradores individuais).
    Cada registro na tabela fretado já contém os dados do colaborador.
    """
    # Data padrão: hoje
    data_hoje = date.today().strftime('%Y-%m-%d')
    data_filtro = request.args.get('data_filtro', data_hoje)
    empresa_id = request.args.get('empresa_id', '')
    planta_id = request.args.get('planta_id', '')
    bloco_id = request.args.get('bloco_id', '')

    # Query base - NÃO PRECISA DE JOIN!
    # A tabela fretado JÁ TEM todos os dados do colaborador
    query = Fretado.query

    # Filtro por data
    if data_filtro:
        try:
            data_obj = datetime.strptime(data_filtro, '%Y-%m-%d').date()

            # Filtra por qualquer horário (entrada, saída ou desligamento)
            # Usa func.date() para comparar apenas a data
            query = query.filter(
                or_(
                    and_(
                        Fretado.horario_entrada.isnot(None),
                        func.date(Fretado.horario_entrada) == data_obj
                    ),
                    and_(
                        Fretado.horario_saida.isnot(None),
                        func.date(Fretado.horario_saida) == data_obj
                    ),
                    and_(
                        Fretado.horario_desligamento.isnot(None),
                        func.date(Fretado.horario_desligamento) == data_obj
                    )
                )
            )
        except ValueError:
            flash('Data inválida', 'error')

    # Filtro por empresa
    if empresa_id:
        try:
            query = query.filter(Fretado.empresa_id == int(empresa_id))
        except ValueError:
            pass

    # Filtro por planta
    if planta_id:
        try:
            query = query.filter(Fretado.planta_id == int(planta_id))
        except ValueError:
            pass

    # Filtro por bloco
    if bloco_id:
        try:
            query = query.filter(Fretado.bloco_id == int(bloco_id))
        except ValueError:
            pass

    # Filtro por permissão do usuário
    if current_user.role == 'gerente':
        gerente = Gerente.query.filter_by(user_id=current_user.id).first()
        if gerente:
            query = query.filter(Fretado.empresa_id == gerente.empresa_id)
            # Gerente vê fretados de todas as suas plantas
            plantas_ids = [p.id for p in gerente.plantas.all()]
            if plantas_ids:
                query = query.filter(Fretado.planta_id.in_(plantas_ids))

    elif current_user.role == 'supervisor':
        supervisor = Supervisor.query.filter_by(
            user_id=current_user.id).first()
        if supervisor:
            query = query.filter(Fretado.empresa_id == supervisor.empresa_id)
            if supervisor.plantas:
                plantas_ids = [p.id for p in supervisor.plantas]
                query = query.filter(Fretado.planta_id.in_(plantas_ids))

    # Ordena por data/horário (usa coalesce para pegar o primeiro horário disponível)
    fretados = query.order_by(
        func.coalesce(
            Fretado.horario_entrada,
            Fretado.horario_saida,
            Fretado.horario_desligamento
        ).desc()
    ).all()

    # ===================================================================
    # MONTA LISTA DE COLABORADORES
    # ===================================================================
    # A tabela fretado JÁ TEM os dados, só precisa formatar
    colaboradores_fretados = []

    for fretado in fretados:
        # Determina o horário principal
        horario = fretado.horario_entrada or fretado.horario_saida or fretado.horario_desligamento
        horario_str = horario.strftime('%d/%m/%Y %H:%M') if horario else 'N/A'

        colaboradores_fretados.append({
            'fretado_id': fretado.id,
            'matricula': fretado.matricula or 'N/A',
            'nome_colaborador': fretado.nome_colaborador or 'N/A',
            'bairro': fretado.bairro or 'N/A',
            'cidade': fretado.cidade or 'N/A',
            'telefone': fretado.telefone or 'N/A',
            'empresa': fretado.empresa.nome if fretado.empresa else 'N/A',
            'planta': fretado.planta.nome if fretado.planta else 'N/A',
            'bloco': fretado.bloco.codigo_bloco if fretado.bloco else 'N/A',
            'tipo_linha': fretado.tipo_linha or 'N/A',
            'tipo_corrida': fretado.tipo_corrida or 'N/A',
            'data_horario': horario_str,
            'status': fretado.status
        })

    # Busca dados para os filtros
    empresas = Empresa.query.filter_by(status='Ativa').all()
    plantas = Planta.query.all()
    blocos = Bloco.query.filter_by(status='Ativo').all()

    # Filtra plantas por empresa se selecionada
    if empresa_id:
        try:
            plantas = [p for p in plantas if p.empresa_id == int(empresa_id)]
        except ValueError:
            pass

    return render_template(
        'fretados/listagem.html',
        colaboradores_fretados=colaboradores_fretados,
        empresas=empresas,
        plantas=plantas,
        blocos=blocos,
        filtros=request.args,
        data_hoje=data_hoje
    )


@admin_bp.route('/fretados/exportar')
@login_required
@permission_required(['admin', 'gerente', 'supervisor', 'operador'])
def exportar_fretados():
    """
    Exporta os fretados filtrados para Excel ou CSV.
    """
    formato = request.args.get('formato', 'excel')  # 'excel' ou 'csv'

    # Aplica os mesmos filtros da listagem
    data_hoje = date.today().strftime('%Y-%m-%d')
    data_filtro = request.args.get('data_filtro', data_hoje)
    empresa_id = request.args.get('empresa_id', '')
    planta_id = request.args.get('planta_id', '')
    bloco_id = request.args.get('bloco_id', '')

    # Query base
    query = Fretado.query

    # Filtro por data
    if data_filtro:
        try:
            data_obj = datetime.strptime(data_filtro, '%Y-%m-%d').date()
            query = query.filter(
                or_(
                    and_(
                        Fretado.horario_entrada.isnot(None),
                        func.date(Fretado.horario_entrada) == data_obj
                    ),
                    and_(
                        Fretado.horario_saida.isnot(None),
                        func.date(Fretado.horario_saida) == data_obj
                    ),
                    and_(
                        Fretado.horario_desligamento.isnot(None),
                        func.date(Fretado.horario_desligamento) == data_obj
                    )
                )
            )
        except ValueError:
            pass

    # Filtro por empresa
    if empresa_id:
        try:
            query = query.filter(Fretado.empresa_id == int(empresa_id))
        except ValueError:
            pass

    # Filtro por planta
    if planta_id:
        try:
            query = query.filter(Fretado.planta_id == int(planta_id))
        except ValueError:
            pass

    # Filtro por bloco
    if bloco_id:
        try:
            query = query.filter(Fretado.bloco_id == int(bloco_id))
        except ValueError:
            pass

    # Filtro por permissão do usuário
    if current_user.role == 'gerente':
        gerente = Gerente.query.filter_by(user_id=current_user.id).first()
        if gerente:
            query = query.filter(Fretado.empresa_id == gerente.empresa_id)
            # Gerente vê fretados de todas as suas plantas
            plantas_ids = [p.id for p in gerente.plantas.all()]
            if plantas_ids:
                query = query.filter(Fretado.planta_id.in_(plantas_ids))

    elif current_user.role == 'supervisor':
        supervisor = Supervisor.query.filter_by(
            user_id=current_user.id).first()
        if supervisor:
            query = query.filter(Fretado.empresa_id == supervisor.empresa_id)
            if supervisor.plantas:
                plantas_ids = [p.id for p in supervisor.plantas]
                query = query.filter(Fretado.planta_id.in_(plantas_ids))

    fretados = query.order_by(
        func.coalesce(
            Fretado.horario_entrada,
            Fretado.horario_saida,
            Fretado.horario_desligamento
        ).desc()
    ).all()

    # Monta lista de colaboradores
    colaboradores_fretados = []

    for fretado in fretados:
        horario = fretado.horario_entrada or fretado.horario_saida or fretado.horario_desligamento
        horario_str = horario.strftime('%d/%m/%Y %H:%M') if horario else 'N/A'

        colaboradores_fretados.append({
            'fretado_id': fretado.id,
            'matricula': fretado.matricula or 'N/A',
            'nome_colaborador': fretado.nome_colaborador or 'N/A',
            'bairro': fretado.bairro or 'N/A',
            'cidade': fretado.cidade or 'N/A',
            'telefone': fretado.telefone or 'N/A',
            'empresa': fretado.empresa.nome if fretado.empresa else 'N/A',
            'planta': fretado.planta.nome if fretado.planta else 'N/A',
            'bloco': fretado.bloco.codigo_bloco if fretado.bloco else 'N/A',
            'tipo_linha': fretado.tipo_linha or 'N/A',
            'tipo_corrida': fretado.tipo_corrida or 'N/A',
            'data_horario': horario_str,
            'status': fretado.status
        })

    if formato == 'csv':
        return exportar_csv(colaboradores_fretados)
    else:
        return exportar_excel(colaboradores_fretados)


def exportar_csv(colaboradores_fretados):
    """Exporta colaboradores em fretados para CSV com separador ;"""
    output = StringIO()
    writer = csv.writer(output, delimiter=';')

    # Cabeçalho
    writer.writerow([
        'ID Fretado',
        'Matrícula',
        'Nome Colaborador',
        'Bairro',
        'Cidade',
        'Telefone',
        'Empresa',
        'Planta',
        'Bloco',
        'Tipo Linha',
        'Tipo Corrida',
        'Data/Horário',
        'Status'
    ])

    # Dados
    for item in colaboradores_fretados:
        writer.writerow([
            item['fretado_id'],
            item['matricula'],
            item['nome_colaborador'],
            item['bairro'],
            item['cidade'],
            item['telefone'],
            item['empresa'],
            item['planta'],
            item['bloco'],
            item['tipo_linha'],
            item['tipo_corrida'],
            item['data_horario'],
            item['status']
        ])

    # Prepara o response
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename=fretados_colaboradores_{date.today().strftime("%Y%m%d")}.csv',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )


def exportar_excel(colaboradores_fretados):
    """Exporta colaboradores em fretados para Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fretados - Colaboradores'

    # Estilos
    header_fill = PatternFill(start_color='4472C4',
                              end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalhos
    headers = [
        'ID Fretado',
        'Matrícula',
        'Nome Colaborador',
        'Bairro',
        'Cidade',
        'Telefone',
        'Empresa',
        'Planta',
        'Bloco',
        'Tipo Linha',
        'Tipo Corrida',
        'Data/Horário',
        'Status'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Dados
    for row_num, item in enumerate(colaboradores_fretados, 2):
        ws.cell(row=row_num, column=1,
                value=item['fretado_id']).border = border
        ws.cell(row=row_num, column=2, value=item['matricula']).border = border
        ws.cell(row=row_num, column=3,
                value=item['nome_colaborador']).border = border
        ws.cell(row=row_num, column=4, value=item['bairro']).border = border
        ws.cell(row=row_num, column=5, value=item['cidade']).border = border
        ws.cell(row=row_num, column=6, value=item['telefone']).border = border
        ws.cell(row=row_num, column=7, value=item['empresa']).border = border
        ws.cell(row=row_num, column=8, value=item['planta']).border = border
        ws.cell(row=row_num, column=9, value=item['bloco']).border = border
        ws.cell(row=row_num, column=10,
                value=item['tipo_linha']).border = border
        ws.cell(row=row_num, column=11,
                value=item['tipo_corrida']).border = border
        ws.cell(row=row_num, column=12,
                value=item['data_horario']).border = border
        ws.cell(row=row_num, column=13, value=item['status']).border = border

    # Ajusta largura das colunas
    column_widths = [12, 12, 25, 20, 15, 15, 20, 20, 12, 12, 15, 18, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(
            col_num)].width = width

    # Salva em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=fretados_colaboradores_{date.today().strftime("%Y%m%d")}.xlsx'
        }
    )
