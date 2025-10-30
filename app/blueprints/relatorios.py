"""
Blueprint de Relatórios - Go Mobi
Versão CORRIGIDA - Com Permissões por Perfil
Data: 14 de Outubro de 2025

Relatórios implementados:
1. Listagem de Solicitações (com permissões)
2. Conferência de Viagens (com permissões e filtros)
3. Conferência de Motoristas (com permissões)

CORREÇÕES IMPLEMENTADAS:
- Permissões por perfil (Admin/Gerente/Supervisor/Motorista)
- Filtros dinâmicos de Planta baseado em Empresa
- Tipo Linha corrigido (FIXA/EXTRA)
- Filtros de Motorista e Colaborador em Conferência de Viagens
- Data da viagem ao invés de data de aceite
- Campo codigo_bloco ao invés de nome_bloco
- Coluna Solicitante em Listagem de Solicitações
- Coluna Colaboradores em Conferência de Motoristas
- Valor Repasse removido de Conferência de Viagens
"""

from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
from datetime import datetime
from io import BytesIO
import openpyxl
import json
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch

from app.models import db, Solicitacao, Viagem, Motorista, Colaborador, Empresa, Planta, Bloco, Supervisor, Gerente

# Criar blueprint
relatorios_bp = Blueprint('relatorios', __name__, url_prefix='/relatorios')


# ========== FUNÇÕES AUXILIARES DE PERMISSÕES ==========

def get_user_empresa_id():
    """Retorna o empresa_id do usuário logado (None se for admin)"""
    if current_user.role == 'admin':
        return None
    elif current_user.role == 'gerente' and current_user.gerente:
        return current_user.gerente.empresa_id
    elif current_user.role == 'supervisor' and current_user.supervisor:
        return current_user.supervisor.empresa_id
    return None


def get_user_planta_id():
    """Retorna o planta_id do usuário logado (None se for admin)"""
    if current_user.role == 'admin':
        return None
    elif current_user.role == 'gerente' and current_user.gerente:
        # Gerente tem múltiplas plantas, retorna None
        return None
    elif current_user.role == 'supervisor' and current_user.supervisor:
        return current_user.supervisor.planta_id
    return None


def get_user_supervisor_id():
    """Retorna o ID do supervisor logado (None se não for supervisor)"""
    if current_user.role == 'supervisor' and current_user.supervisor:
        return current_user.supervisor.id
    return None


def get_user_motorista_id():
    """Retorna o ID do motorista logado (None se não for motorista)"""
    if current_user.role == 'motorista' and current_user.motorista:
        return current_user.motorista.id
    return None


# ========== RELATÓRIO 1: LISTAGEM DE SOLICITAÇÕES ==========

@relatorios_bp.route('/solicitacoes')
@login_required
def listagem_solicitacoes():
    """Exibe a tela de filtros para o relatório de solicitações"""

    # Buscar dados para os filtros baseado no perfil
    empresa_id_usuario = get_user_empresa_id()
    planta_id_usuario = get_user_planta_id()
    supervisor_id_usuario = get_user_supervisor_id()

    # Filtrar empresas baseado no perfil
    if empresa_id_usuario:
        empresas = Empresa.query.filter_by(id=empresa_id_usuario).all()
    else:
        empresas = Empresa.query.all()

    # Filtrar plantas baseado no perfil
    if planta_id_usuario:
        plantas = Planta.query.filter_by(id=planta_id_usuario).all()
    elif empresa_id_usuario:
        plantas = Planta.query.filter_by(empresa_id=empresa_id_usuario).all()
    else:
        plantas = Planta.query.all()

    blocos = Bloco.query.all()

    # Buscar supervisores (apenas para admin)
    if current_user.role == 'admin':
        supervisores = Supervisor.query.all()
    else:
        supervisores = []

    return render_template(
        'relatorios/listagem_solicitacoes.html',
        empresas=empresas,
        plantas=plantas,
        blocos=blocos,
        supervisores=supervisores,
        user_role=current_user.role,
        empresa_id_usuario=empresa_id_usuario,
        planta_id_usuario=planta_id_usuario,
        supervisor_id_usuario=supervisor_id_usuario
    )


@relatorios_bp.route('/solicitacoes/dados', methods=['POST'])
@login_required
def dados_listagem_solicitacoes():
    """Retorna os dados do relatório de solicitações em JSON"""

    try:
        # Receber filtros
        data_inicio = request.form.get('data_inicio')
        data_fim = request.form.get('data_fim')
        empresa_id = request.form.get('empresa_id')
        planta_id = request.form.get('planta_id')
        bloco_id = request.form.get('bloco_id')
        status = request.form.get('status')
        tipo_corrida = request.form.get('tipo_corrida')
        tipo_linha = request.form.get('tipo_linha')
        supervisor_id = request.form.get('supervisor_id')
        horario_filtro = request.form.get('horario')  # NOVO: Filtro de horário

        # Construir query
        query = Solicitacao.query

        # APLICAR PERMISSÕES POR PERFIL
        if current_user.role == 'supervisor':
            # Supervisor vê apenas suas solicitações
            supervisor_id_usuario = get_user_supervisor_id()
            if supervisor_id_usuario:
                query = query.filter_by(supervisor_id=supervisor_id_usuario)

        elif current_user.role == 'gerente':
            # Gerente vê todas as solicitações da planta
            planta_id_usuario = get_user_planta_id()
            if planta_id_usuario:
                query = query.filter_by(planta_id=planta_id_usuario)

        elif current_user.role == 'admin':
            # Admin pode filtrar por supervisor se quiser
            if supervisor_id:
                query = query.filter_by(supervisor_id=supervisor_id)

        # Aplicar filtros de empresa e planta (respeitando permissões)
        if empresa_id:
            # Validar se o usuário tem permissão para ver essa empresa
            empresa_id_usuario = get_user_empresa_id()
            if empresa_id_usuario and int(empresa_id) != empresa_id_usuario:
                return jsonify({
                    'success': False,
                    'message': 'Você não tem permissão para visualizar dados desta empresa'
                }), 403
            query = query.filter_by(empresa_id=empresa_id)
        else:
            # Se não especificou empresa, aplicar filtro do usuário
            empresa_id_usuario = get_user_empresa_id()
            if empresa_id_usuario:
                query = query.filter_by(empresa_id=empresa_id_usuario)

        # Aplicar filtros de data
        if data_inicio and data_fim:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Solicitacao.data_criacao >= data_inicio_dt)
            query = query.filter(Solicitacao.data_criacao <= data_fim_dt)

        if planta_id:
            query = query.filter_by(planta_id=planta_id)

        if bloco_id:
            query = query.filter_by(bloco_id=bloco_id)

        if status:
            query = query.filter_by(status=status)

        if tipo_corrida:
            query = query.filter_by(tipo_corrida=tipo_corrida)

        if tipo_linha:
            query = query.filter_by(tipo_linha=tipo_linha)

        # NOVO: Filtro de horário específico
        if horario_filtro:
            from sqlalchemy import or_, func, cast, Time
            horario_time = datetime.strptime(horario_filtro, '%H:%M').time()
            query = query.filter(
                or_(
                    cast(Solicitacao.horario_entrada, Time) == horario_time,
                    cast(Solicitacao.horario_saida, Time) == horario_time,
                    cast(Solicitacao.horario_desligamento, Time) == horario_time
                )
            )

        # Executar query
        solicitacoes = query.order_by(Solicitacao.data_criacao.desc()).all()

        # Formatar dados para JSON
        dados = []

        for sol in solicitacoes:
            # Buscar dados relacionados
            colaborador = Colaborador.query.get(
                sol.colaborador_id) if sol.colaborador_id else None
            empresa = Empresa.query.get(
                sol.empresa_id) if sol.empresa_id else None
            planta = Planta.query.get(sol.planta_id) if sol.planta_id else None
            bloco = Bloco.query.get(sol.bloco_id) if sol.bloco_id else None
            supervisor = Supervisor.query.get(
                sol.supervisor_id) if sol.supervisor_id else None

            dados.append({
                'id': sol.id,
                'data_criacao': sol.data_criacao.strftime('%d/%m/%Y %H:%M') if sol.data_criacao else '',
                'colaborador': colaborador.nome if colaborador else 'N/A',
                'empresa': empresa.nome if empresa else 'N/A',
                'planta': planta.nome if planta else 'N/A',
                'bloco': bloco.codigo_bloco if bloco else 'N/A',  # CORRIGIDO: usar codigo_bloco
                'tipo_linha': sol.tipo_linha or 'N/A',
                'tipo_corrida': sol.tipo_corrida or 'N/A',
                'status': sol.status or 'N/A',
                # ADICIONADO: nome do solicitante
                'solicitante': supervisor.nome if supervisor else 'N/A',
                'horario_entrada': sol.horario_entrada.strftime('%H:%M') if sol.horario_entrada else '',
                'horario_saida': sol.horario_saida.strftime('%H:%M') if sol.horario_saida else '',
                'horario_desligamento': sol.horario_desligamento.strftime('%H:%M') if sol.horario_desligamento else ''
            })

        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados)
        })

    except Exception as e:
        current_app.logger.error(
            f"Erro ao gerar relatório de solicitações: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar relatório: {str(e)}'
        }), 500


# ========== RELATÓRIO 2: CONFERÊNCIA DE VIAGENS ==========

@relatorios_bp.route('/conferencia-viagens')
@login_required
def conferencia_viagens():
    """Exibe a tela de filtros para o relatório de conferência de viagens"""

    # Buscar dados para os filtros baseado no perfil
    empresa_id_usuario = get_user_empresa_id()
    planta_id_usuario = get_user_planta_id()

    # Filtrar empresas baseado no perfil
    if empresa_id_usuario:
        empresas = Empresa.query.filter_by(id=empresa_id_usuario).all()
    else:
        empresas = Empresa.query.all()

    # Filtrar plantas baseado no perfil
    if planta_id_usuario:
        plantas = Planta.query.filter_by(id=planta_id_usuario).all()
    elif empresa_id_usuario:
        plantas = Planta.query.filter_by(empresa_id=empresa_id_usuario).all()
    else:
        plantas = Planta.query.all()

    # Buscar motoristas e colaboradores para filtros
    motoristas = Motorista.query.all()
    colaboradores = Colaborador.query.all()

    return render_template(
        'relatorios/conferencia_viagens.html',
        empresas=empresas,
        plantas=plantas,
        motoristas=motoristas,
        colaboradores=colaboradores,
        user_role=current_user.role,
        empresa_id_usuario=empresa_id_usuario,
        planta_id_usuario=planta_id_usuario
    )


@relatorios_bp.route('/conferencia-viagens/dados', methods=['POST'])
@login_required
def dados_conferencia_viagens():
    """Retorna os dados do relatório de conferência de viagens em JSON"""

    try:
        # Receber filtros
        data_inicio = request.form.get('data_inicio')
        data_fim = request.form.get('data_fim')
        empresa_id = request.form.get('empresa_id')
        planta_id = request.form.get('planta_id')
        status = request.form.get('status')
        tipo_corrida = request.form.get('tipo_corrida')
        tipo_linha = request.form.get('tipo_linha')
        motorista_id = request.form.get('motorista_id')  # NOVO FILTRO
        colaborador_id = request.form.get('colaborador_id')  # NOVO FILTRO

        # Construir query
        query = Viagem.query.options(db.joinedload(Viagem.hora_parada))

        # APLICAR PERMISSÕES POR PERFIL
        if current_user.role == 'supervisor':
            # Supervisor vê apenas viagens das solicitações que ele fez
            supervisor_id_usuario = get_user_supervisor_id()
            if supervisor_id_usuario:
                # Buscar IDs de viagens das solicitações do supervisor
                solicitacoes_supervisor = Solicitacao.query.filter_by(
                    supervisor_id=supervisor_id_usuario).all()
                viagem_ids = [
                    sol.viagem_id for sol in solicitacoes_supervisor if sol.viagem_id]
                if viagem_ids:
                    query = query.filter(Viagem.id.in_(viagem_ids))
                else:
                    # Se não tem viagens, retorna vazio
                    return jsonify({
                        'success': True,
                        'dados': [],
                        'total': 0,
                        'valor_total': 0.0
                    })

        elif current_user.role == 'gerente':
            # Gerente vê todas as viagens da planta
            planta_id_usuario = get_user_planta_id()
            if planta_id_usuario:
                query = query.filter_by(planta_id=planta_id_usuario)

        # Aplicar filtros de empresa e planta (respeitando permissões)
        if empresa_id:
            # Validar se o usuário tem permissão para ver essa empresa
            empresa_id_usuario = get_user_empresa_id()
            if empresa_id_usuario and int(empresa_id) != empresa_id_usuario:
                return jsonify({
                    'success': False,
                    'message': 'Você não tem permissão para visualizar dados desta empresa'
                }), 403
            query = query.filter_by(empresa_id=empresa_id)
        else:
            # Se não especificou empresa, aplicar filtro do usuário
            empresa_id_usuario = get_user_empresa_id()
            if empresa_id_usuario:
                query = query.filter_by(empresa_id=empresa_id_usuario)

        # CORRIGIDO: Filtrar por data da viagem (horario_entrada, horario_saida, horario_desligamento)
        if data_inicio and data_fim:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)

            # Filtrar por qualquer um dos horários da viagem
            query = query.filter(
                db.or_(
                    db.and_(Viagem.horario_entrada >= data_inicio_dt,
                            Viagem.horario_entrada <= data_fim_dt),
                    db.and_(Viagem.horario_saida >= data_inicio_dt,
                            Viagem.horario_saida <= data_fim_dt),
                    db.and_(Viagem.horario_desligamento >= data_inicio_dt,
                            Viagem.horario_desligamento <= data_fim_dt)
                )
            )

        if planta_id:
            query = query.filter_by(planta_id=planta_id)

        if status:
            query = query.filter_by(status=status)

        if tipo_corrida:
            query = query.filter_by(tipo_corrida=tipo_corrida)

        if tipo_linha:
            query = query.filter_by(tipo_linha=tipo_linha)

        # NOVO: Filtro por motorista
        if motorista_id:
            query = query.filter_by(motorista_id=motorista_id)

        # NOVO: Filtro por colaborador (busca em colaboradores_ids)
        if colaborador_id:
            # Precisa buscar viagens que contenham esse colaborador_id no JSON
            query = query.filter(
                Viagem.colaboradores_ids.like(f'%{colaborador_id}%'))

        # Executar query
        viagens = query.order_by(Viagem.id.desc()).all()

        # Formatar dados para JSON
        dados = []
        valor_total = 0.0

        for viagem in viagens:
            # Buscar dados relacionados
            empresa = Empresa.query.get(
                viagem.empresa_id) if viagem.empresa_id else None
            planta = Planta.query.get(
                viagem.planta_id) if viagem.planta_id else None
            bloco = Bloco.query.get(
                viagem.bloco_id) if viagem.bloco_id else None
            motorista = Motorista.query.get(
                viagem.motorista_id) if viagem.motorista_id else None

            # Buscar colaboradores
            colaboradores = []
            if viagem.colaboradores_ids:
                try:
                    col_ids = json.loads(viagem.colaboradores_ids)
                    for col_id in col_ids:
                        col = Colaborador.query.get(col_id)
                        if col:
                            colaboradores.append(col.nome)
                except:
                    pass

            # Calcular valor (incluindo hora parada)
            valor_viagem = float(viagem.valor) if viagem.valor else 0.0
            if viagem.hora_parada:
                valor_viagem += float(viagem.hora_parada.valor_adicional)
            valor_total += valor_viagem

            # Determinar data da viagem (prioridade: entrada > saida > desligamento)
            data_viagem = None
            if viagem.horario_entrada:
                data_viagem = viagem.horario_entrada
            elif viagem.horario_saida:
                data_viagem = viagem.horario_saida
            elif viagem.horario_desligamento:
                data_viagem = viagem.horario_desligamento

            dados.append({
                'id': viagem.id,
                # CORRIGIDO
                'data_viagem': data_viagem.strftime('%d/%m/%Y') if data_viagem else 'N/A',
                'empresa': empresa.nome if empresa else 'N/A',
                'planta': planta.nome if planta else 'N/A',
                'bloco': bloco.codigo_bloco if bloco else 'N/A',  # CORRIGIDO: usar codigo_bloco
                'tipo_linha': viagem.tipo_linha or 'N/A',
                'tipo_corrida': viagem.tipo_corrida or 'N/A',
                'status': viagem.status or 'N/A',
                'motorista': motorista.nome if motorista else viagem.nome_motorista or 'N/A',
                'placa': viagem.placa_veiculo or 'N/A',
                'colaboradores': ', '.join(colaboradores) if colaboradores else 'N/A',
                'qtd_passageiros': viagem.quantidade_passageiros or 0,
                'valor': valor_viagem,
                'horario_entrada': viagem.horario_entrada.strftime('%H:%M') if viagem.horario_entrada else '',
                'horario_saida': viagem.horario_saida.strftime('%H:%M') if viagem.horario_saida else '',
                'horario_desligamento': viagem.horario_desligamento.strftime('%H:%M') if viagem.horario_desligamento else ''
            })

        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados),
            'valor_total': valor_total
        })

    except Exception as e:
        current_app.logger.error(
            f"Erro ao gerar relatório de viagens: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar relatório: {str(e)}'
        }), 500


# ========== RELATÓRIO 3: CONFERÊNCIA DE MOTORISTAS ==========

@relatorios_bp.route('/conferencia-motoristas')
@login_required
def conferencia_motoristas():
    """Exibe a tela de filtros para o relatório de conferência de motoristas"""

    motorista_id_usuario = get_user_motorista_id()

    # Se for motorista, só pode ver suas próprias viagens
    if current_user.role == 'motorista':
        motoristas = Motorista.query.filter_by(id=motorista_id_usuario).all()
    else:
        motoristas = Motorista.query.all()

    return render_template(
        'relatorios/conferencia_motoristas.html',
        motoristas=motoristas,
        user_role=current_user.role,
        motorista_id_usuario=motorista_id_usuario
    )


@relatorios_bp.route('/conferencia-motoristas/dados', methods=['POST'])
@login_required
def dados_conferencia_motoristas():
    """Retorna os dados do relatório de conferência de motoristas em JSON"""

    try:
        # Receber filtros
        data_inicio = request.form.get('data_inicio')
        data_fim = request.form.get('data_fim')
        motorista_id = request.form.get('motorista_id')
        status = request.form.get('status')

        # Construir query
        query = Viagem.query.options(db.joinedload(Viagem.hora_parada))

        # APLICAR PERMISSÕES POR PERFIL
        motorista_id_usuario = get_user_motorista_id()

        if current_user.role == 'motorista':
            # Motorista só vê suas próprias viagens
            if motorista_id_usuario:
                query = query.filter_by(motorista_id=motorista_id_usuario)
            else:
                # Se não tem motorista_id, retorna vazio
                return jsonify({
                    'success': True,
                    'dados': [],
                    'total': 0,
                    'valor_total_repasse': 0.0
                })
        else:
            # Admin pode filtrar por qualquer motorista
            if motorista_id:
                query = query.filter_by(motorista_id=motorista_id)

        # CORRIGIDO: Filtrar por data da viagem (não por data_inicio ou data_finalizacao)
        if data_inicio and data_fim:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)

            # Filtrar por qualquer um dos horários da viagem
            query = query.filter(
                db.or_(
                    db.and_(Viagem.horario_entrada >= data_inicio_dt,
                            Viagem.horario_entrada <= data_fim_dt),
                    db.and_(Viagem.horario_saida >= data_inicio_dt,
                            Viagem.horario_saida <= data_fim_dt),
                    db.and_(Viagem.horario_desligamento >= data_inicio_dt,
                            Viagem.horario_desligamento <= data_fim_dt)
                )
            )

        if status:
            query = query.filter_by(status=status)

        # Executar query
        viagens = query.order_by(Viagem.id.desc()).all()

        # Formatar dados para JSON
        dados = []
        valor_total_repasse = 0.0

        for viagem in viagens:
            motorista = Motorista.query.get(
                viagem.motorista_id) if viagem.motorista_id else None
            empresa = Empresa.query.get(
                viagem.empresa_id) if viagem.empresa_id else None
            planta = Planta.query.get(
                viagem.planta_id) if viagem.planta_id else None

            # Buscar colaboradores e bairros da viagem
            colaboradores = []
            bairros = []
            if viagem.colaboradores_ids:
                try:
                    col_ids = json.loads(viagem.colaboradores_ids)
                    for col_id in col_ids:
                        col = Colaborador.query.get(col_id)
                        if col:
                            colaboradores.append(col.nome)
                            # Buscar bairro diretamente do colaborador
                            if col.bairro:
                                bairros.append(col.bairro)
                            else:
                                bairros.append('N/A')
                except:
                    pass

            # Calcular repasse (incluindo hora parada)
            valor_repasse = float(
                viagem.valor_repasse) if viagem.valor_repasse else 0.0
            if viagem.hora_parada:
                valor_repasse += float(viagem.hora_parada.repasse_adicional)
            valor_total_repasse += valor_repasse

            # Determinar data da viagem
            data_viagem = None
            if viagem.horario_entrada:
                data_viagem = viagem.horario_entrada
            elif viagem.horario_saida:
                data_viagem = viagem.horario_saida
            elif viagem.horario_desligamento:
                data_viagem = viagem.horario_desligamento

            dados.append({
                'id': viagem.id,
                # CORRIGIDO
                'data_viagem': data_viagem.strftime('%d/%m/%Y') if data_viagem else 'N/A',
                'horario': data_viagem.strftime('%H:%M') if data_viagem else 'N/A',
                'motorista': motorista.nome if motorista else viagem.nome_motorista or 'N/A',
                'empresa': empresa.nome if empresa else 'N/A',
                'planta': planta.nome if planta else 'N/A',
                'tipo_corrida': viagem.tipo_corrida or 'N/A',
                'status': viagem.status or 'N/A',
                'placa': viagem.placa_veiculo or 'N/A',
                'qtd_passageiros': viagem.quantidade_passageiros or 0,
                'colaboradores': '\n'.join(colaboradores) if colaboradores else 'N/A',
                'bairros': '\n'.join(bairros) if bairros else 'N/A',
                'valor_repasse': valor_repasse
            })

        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados),
            'valor_total_repasse': valor_total_repasse
        })

    except Exception as e:
        current_app.logger.error(
            f"Erro ao gerar relatório de motoristas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar relatório: {str(e)}'
        }), 500


# ========== ROTA AUXILIAR - PLANTAS POR EMPRESA ==========

@relatorios_bp.route('/plantas-por-empresa/<int:empresa_id>')
@login_required
def plantas_por_empresa(empresa_id):
    """Retorna as plantas de uma empresa específica (para filtro dinâmico)"""
    try:
        plantas = Planta.query.filter_by(empresa_id=empresa_id).all()
        return jsonify({
            'success': True,
            'plantas': [{'id': p.id, 'nome': p.nome} for p in plantas]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ========== EXPORTAÇÃO EXCEL ==========

@relatorios_bp.route('/exportar-excel/<tipo>', methods=['POST'])
@login_required
def exportar_excel(tipo):
    """Exporta o relatório para Excel"""
    try:
        dados_json = request.form.get('dados')
        dados = json.loads(dados_json)

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Definir colunas baseado no tipo
        if tipo == 'solicitacoes':
            ws.title = "Listagem de Solicitações"
            colunas = ['ID', 'Data Solicitação', 'Colaborador', 'Empresa', 'Planta', 'Bloco',
                       'Tipo Linha', 'Tipo Corrida', 'Horário', 'Solicitante', 'Status']
            campos = ['id', 'data_criacao', 'colaborador', 'empresa', 'planta', 'bloco',
                      'tipo_linha', 'tipo_corrida', 'horario', 'solicitante', 'status']

        elif tipo == 'viagens':
            ws.title = "Conferência de Viagens"
            colunas = ['ID', 'Data', 'Horário', 'Empresa', 'Planta', 'Bloco', 'Tipo Linha',
                       'Tipo Corrida', 'Motorista', 'Passageiros', 'Colaboradores',
                       'Valor', 'Status']
            campos = ['id', 'data_viagem', 'horario', 'empresa', 'planta', 'bloco', 'tipo_linha',
                      'tipo_corrida', 'motorista', 'qtd_passageiros', 'colaboradores',
                      'valor', 'status']

        elif tipo == 'motoristas':
            ws.title = "Conferência de Motoristas"
            colunas = ['ID', 'Data', 'Horário', 'Motorista', 'Empresa', 'Planta',
                       'Tipo Corrida', 'Status', 'Placa', 'Passageiros', 'Colaboradores',
                       'Valor Repasse']
            campos = ['id', 'data_viagem', 'horario', 'motorista', 'empresa', 'planta',
                      'tipo_corrida', 'status', 'placa', 'qtd_passageiros', 'colaboradores',
                      'valor_repasse']

        # Escrever cabeçalho
        for col_num, coluna in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Escrever dados
        for row_num, item in enumerate(dados, 2):
            for col_num, campo in enumerate(campos, 1):
                # Para solicitações e viagens, calcular horário unificado
                if campo == 'horario' and tipo in ['solicitacoes', 'viagens']:
                    tipo_corrida = item.get('tipo_corrida', '')
                    if tipo_corrida == 'entrada':
                        valor = item.get('horario_entrada', '')
                    elif tipo_corrida == 'saida':
                        valor = item.get('horario_saida', '')
                    elif tipo_corrida == 'desligamento':
                        valor = item.get('horario_desligamento', '')
                    else:
                        valor = ''
                else:
                    valor = item.get(campo, '')
                    if isinstance(valor, float):
                        valor = f"R$ {valor:.2f}"
                ws.cell(row=row_num, column=col_num, value=valor)

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar arquivo
        filename = f"relatorio_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"Erro ao exportar Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ========== EXPORTAÇÃO PDF ==========

@relatorios_bp.route('/exportar-pdf/<tipo>', methods=['POST'])
@login_required
def exportar_pdf(tipo):
    """Exporta o relatório para PDF"""
    try:
        dados_json = request.form.get('dados')
        dados = json.loads(dados_json)

        # Criar PDF em memória
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(
            A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']

        # Definir colunas e dados baseado no tipo
        if tipo == 'solicitacoes':
            titulo = "Listagem de Solicitações"
            colunas = ['ID', 'Data', 'Colaborador', 'Empresa', 'Planta', 'Bloco',
                       'Tipo Linha', 'Tipo Corrida', 'Horário', 'Solicitante', 'Status']
            campos = ['id', 'data_criacao', 'colaborador', 'empresa', 'planta', 'bloco',
                      'tipo_linha', 'tipo_corrida', 'horario', 'solicitante', 'status']
            # Larguras ajustadas para A4 landscape (11 polegadas disponíveis)
            col_widths = [0.35*inch, 1*inch, 1.8*inch, 0.9*inch, 0.9*inch, 0.65*inch,
                          0.7*inch, 0.85*inch, 0.6*inch, 1.1*inch, 0.85*inch]

        elif tipo == 'viagens':
            titulo = "Conferência de Viagens"
            colunas = ['ID', 'Data', 'Hor.', 'Empresa', 'Planta', 'Bloco', 'Tipo Linha',
                       'Tipo Corrida', 'Motorista', 'Pass.', 'Colaboradores', 'Valor', 'Status']
            campos = ['id', 'data_viagem', 'horario', 'empresa', 'planta', 'bloco', 'tipo_linha',
                      'tipo_corrida', 'motorista', 'qtd_passageiros', 'colaboradores', 'valor', 'status']
            # Larguras ajustadas para A4 landscape (13 colunas, SEM Placa)
            col_widths = [0.3*inch, 0.8*inch, 0.5*inch, 0.9*inch, 0.9*inch, 0.6*inch, 0.65*inch,
                          0.8*inch, 1.3*inch, 0.4*inch, 1.8*inch, 0.85*inch, 0.75*inch]

        elif tipo == 'motoristas':
            titulo = "Conferência de Motoristas"
            colunas = ['ID', 'Data', 'Hor.', 'Motorista', 'Empresa', 'Planta', 'Tipo Corrida',
                       'Status', 'Placa', 'Pass.', 'Colaboradores', 'Bairros', 'Valor Repasse']
            campos = ['id', 'data_viagem', 'horario', 'motorista', 'empresa', 'planta', 'tipo_corrida',
                      'status', 'placa', 'qtd_passageiros', 'colaboradores', 'bairros', 'valor_repasse']
            col_widths = [0.3*inch, 0.65*inch, 0.45*inch, 1.1*inch, 0.85*inch, 0.75*inch, 0.7*inch,
                          0.65*inch, 0.6*inch, 0.35*inch, 1.5*inch, 1.2*inch, 0.75*inch]

        # Adicionar título
        elements.append(Paragraph(titulo, title_style))
        elements.append(Paragraph("<br/>", styles['Normal']))

        # Preparar dados da tabela com Paragraphs para quebra de linha
        table_data = [colunas]
        for item in dados:
            row = []
            for campo in campos:
                # Para solicitações e viagens, calcular horário unificado
                if campo == 'horario' and tipo in ['solicitacoes', 'viagens']:
                    tipo_corrida = item.get('tipo_corrida', '')
                    if tipo_corrida == 'entrada':
                        valor = item.get('horario_entrada', '')
                    elif tipo_corrida == 'saida':
                        valor = item.get('horario_saida', '')
                    elif tipo_corrida == 'desligamento':
                        valor = item.get('horario_desligamento', '')
                    else:
                        valor = ''
                else:
                    valor = item.get(campo, '')
                    if isinstance(valor, float):
                        valor = f"R$ {valor:.2f}"

                # Usar Paragraph para campos de texto longo (quebra automática)
                if campo in ['colaborador', 'solicitante', 'colaboradores', 'bairros', 'motorista']:
                    cell_style = styles['Normal']
                    cell_style.fontSize = 6
                    cell_style.leading = 7
                    cell_style.alignment = 0  # Esquerda
                    row.append(Paragraph(str(valor), cell_style))
                # Campos numéricos: sem quebra de linha
                elif campo in ['valor', 'valor_repasse']:
                    row.append(str(valor))
                else:
                    row.append(str(valor))
            table_data.append(row)

        # Criar tabela com altura de linha automática
        table = Table(table_data, colWidths=col_widths,
                      repeatRows=1, rowHeights=None)

        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            # Dados
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#F2F2F2')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))

        # Alinhamento específico por tipo de relatório
        if tipo == 'solicitacoes':
            # Colaborador (col 2) e Solicitante (col 9) à esquerda
            table.setStyle(TableStyle([
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                ('ALIGN', (9, 1), (9, -1), 'LEFT'),
            ]))
        elif tipo == 'viagens':
            # Motorista (col 8) e Colaboradores (col 10) à esquerda, Valor (col 11) à direita
            table.setStyle(TableStyle([
                ('ALIGN', (8, 1), (8, -1), 'LEFT'),
                ('ALIGN', (10, 1), (10, -1), 'LEFT'),
                ('ALIGN', (11, 1), (11, -1), 'RIGHT'),
            ]))
        elif tipo == 'motoristas':
            # Motorista (col 3), Colaboradores (col 10) e Bairros (col 11) à esquerda, Valor Repasse (col 12) à direita
            table.setStyle(TableStyle([
                ('ALIGN', (3, 1), (3, -1), 'LEFT'),
                ('ALIGN', (10, 1), (10, -1), 'LEFT'),
                ('ALIGN', (11, 1), (11, -1), 'LEFT'),
                ('ALIGN', (12, 1), (12, -1), 'RIGHT'),
            ]))

        elements.append(table)

        # Gerar PDF
        doc.build(elements)
        output.seek(0)

        # Enviar arquivo
        filename = f"relatorio_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"Erro ao exportar PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
