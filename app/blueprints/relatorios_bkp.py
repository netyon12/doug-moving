"""
Blueprint de Relatórios - Go Mobi
Versão SIMPLIFICADA E FUNCIONAL
Data: 13 de Outubro de 2025

Relatórios implementados:
1. Listagem de Solicitações
2. Conferência de Viagens  
3. Conferência de Motoristas
"""

from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

from app.models import db, Solicitacao, Viagem, Motorista, Colaborador, Empresa, Planta, Bloco

# Criar blueprint
relatorios_bp = Blueprint('relatorios', __name__, url_prefix='/relatorios')


# ========== RELATÓRIO 1: LISTAGEM DE SOLICITAÇÕES ==========

@relatorios_bp.route('/solicitacoes')
@login_required
def listagem_solicitacoes():
    """Exibe a tela de filtros para o relatório de solicitações"""
    
    # Buscar dados para os filtros
    empresas = Empresa.query.all()
    plantas = Planta.query.all()
    blocos = Bloco.query.all()
    
    return render_template(
        'relatorios/listagem_solicitacoes.html',
        empresas=empresas,
        plantas=plantas,
        blocos=blocos
    )


@relatorios_bp.route('/solicitacoes/dados', methods=['POST'])
@login_required
def dados_listagem_solicitacoes():
    """Retorna os dados do relatório de solicitações em JSON"""
    
    try:
        # Receber filtros
        data_inicio = request.form.get('data_inicio')
        data_fim = request.form.get('data_fim')
        empresa_id = request.form.get('empresa_id')
        planta_id = request.form.get('planta_id')
        bloco_id = request.form.get('bloco_id')
        status = request.form.get('status')
        tipo_corrida = request.form.get('tipo_corrida')
        tipo_linha = request.form.get('tipo_linha')
        
        # Construir query
        query = Solicitacao.query
        
        # Aplicar filtros
        if data_inicio and data_fim:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Solicitacao.data_criacao >= data_inicio_dt)
            query = query.filter(Solicitacao.data_criacao <= data_fim_dt)
        
        if empresa_id:
            query = query.filter_by(empresa_id=empresa_id)
        
        if planta_id:
            query = query.filter_by(planta_id=planta_id)
        
        if bloco_id:
            query = query.filter_by(bloco_id=bloco_id)
        
        if status:
            query = query.filter_by(status=status)
        
        if tipo_corrida:
            query = query.filter_by(tipo_corrida=tipo_corrida)
        
        if tipo_linha:
            query = query.filter_by(tipo_linha=tipo_linha)
        
        # Executar query
        solicitacoes = query.order_by(Solicitacao.data_criacao.desc()).all()
        
        # Formatar dados para JSON
        dados = []
        for sol in solicitacoes:
            # Buscar dados relacionados
            colaborador = Colaborador.query.get(sol.colaborador_id)
            empresa = Empresa.query.get(sol.empresa_id)
            planta = Planta.query.get(sol.planta_id)
            bloco = Bloco.query.get(sol.bloco_id)
            
            dados.append({
                'id': sol.id,
                'data_criacao': sol.data_criacao.strftime('%d/%m/%Y %H:%M') if sol.data_criacao else '',
                'colaborador': colaborador.nome if colaborador else 'N/A',
                'empresa': empresa.nome if empresa else 'N/A',
                'planta': planta.nome if planta else 'N/A',
                'bloco': bloco.nome_bloco if bloco else 'N/A',
                'tipo_linha': sol.tipo_linha or 'N/A',
                'tipo_corrida': sol.tipo_corrida or 'N/A',
                'status': sol.status or 'N/A',
                'horario_entrada': sol.horario_entrada.strftime('%H:%M') if sol.horario_entrada else '',
                'horario_saida': sol.horario_saida.strftime('%H:%M') if sol.horario_saida else '',
                'valor': float(sol.valor) if sol.valor else 0.0,
                'valor_repasse': float(sol.valor_repasse) if sol.valor_repasse else 0.0
            })
        
        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados)
        })
        
    except Exception as e:
        current_app.logger.error(f"Erro ao gerar relatório de solicitações: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar relatório: {str(e)}'
        }), 500


# ========== RELATÓRIO 2: CONFERÊNCIA DE VIAGENS ==========

@relatorios_bp.route('/conferencia-viagens')
@login_required
def conferencia_viagens():
    """Exibe a tela de filtros para o relatório de conferência de viagens"""
    
    empresas = Empresa.query.all()
    plantas = Planta.query.all()
    
    return render_template(
        'relatorios/conferencia_viagens.html',
        empresas=empresas,
        plantas=plantas
    )


@relatorios_bp.route('/conferencia-viagens/dados', methods=['POST'])
@login_required
def dados_conferencia_viagens():
    """Retorna os dados do relatório de conferência de viagens em JSON"""
    
    try:
        # Receber filtros
        data_inicio = request.form.get('data_inicio')
        data_fim = request.form.get('data_fim')
        empresa_id = request.form.get('empresa_id')
        planta_id = request.form.get('planta_id')
        status = request.form.get('status')
        tipo_corrida = request.form.get('tipo_corrida')
        tipo_linha = request.form.get('tipo_linha')
        
        # Construir query
        query = Viagem.query
        
        # Aplicar filtros
        if data_inicio and data_fim:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)
            
            # Filtrar por data_inicio ou data_criacao
            query = query.filter(
                db.or_(
                    db.and_(Viagem.data_inicio >= data_inicio_dt, Viagem.data_inicio <= data_fim_dt),
                    db.and_(Viagem.data_criacao >= data_inicio_dt, Viagem.data_criacao <= data_fim_dt)
                )
            )
        
        if empresa_id:
            query = query.filter_by(empresa_id=empresa_id)
        
        if planta_id:
            query = query.filter_by(planta_id=planta_id)
        
        if status:
            query = query.filter_by(status=status)
        
        if tipo_corrida:
            query = query.filter_by(tipo_corrida=tipo_corrida)
        
        if tipo_linha:
            query = query.filter_by(tipo_linha=tipo_linha)
        
        # Executar query
        viagens = query.order_by(Viagem.data_inicio.desc()).all()
        
        # Formatar dados para JSON
        dados = []
        valor_total = 0.0
        
        for viagem in viagens:
            # Buscar dados relacionados
            empresa = Empresa.query.get(viagem.empresa_id) if viagem.empresa_id else None
            planta = Planta.query.get(viagem.planta_id) if viagem.planta_id else None
            bloco = Bloco.query.get(viagem.bloco_id) if viagem.bloco_id else None
            motorista = Motorista.query.get(viagem.motorista_id) if viagem.motorista_id else None
            
            # Buscar colaboradores
            colaboradores = []
            if viagem.colaboradores_ids:
                import json
                try:
                    col_ids = json.loads(viagem.colaboradores_ids)
                    for col_id in col_ids:
                        col = Colaborador.query.get(col_id)
                        if col:
                            colaboradores.append(col.nome)
                except:
                    pass
            
            # Calcular valor
            valor_viagem = float(viagem.valor) if viagem.valor else 0.0
            valor_total += valor_viagem
            
            dados.append({
                'id': viagem.id,
                'data_inicio': viagem.data_inicio.strftime('%d/%m/%Y %H:%M') if viagem.data_inicio else '',
                'empresa': empresa.nome if empresa else 'N/A',
                'planta': planta.nome if planta else 'N/A',
                'bloco': bloco.nome_bloco if bloco else 'N/A',
                'tipo_linha': viagem.tipo_linha or 'N/A',
                'tipo_corrida': viagem.tipo_corrida or 'N/A',
                'status': viagem.status or 'N/A',
                'motorista': motorista.nome if motorista else viagem.nome_motorista or 'N/A',
                'placa': viagem.placa_veiculo or 'N/A',
                'colaboradores': ', '.join(colaboradores) if colaboradores else 'N/A',
                'qtd_passageiros': viagem.quantidade_passageiros or 0,
                'valor': valor_viagem,
                'valor_repasse': float(viagem.valor_repasse) if viagem.valor_repasse else 0.0,
                'horario_entrada': viagem.horario_entrada.strftime('%H:%M') if viagem.horario_entrada else '',
                'horario_saida': viagem.horario_saida.strftime('%H:%M') if viagem.horario_saida else ''
            })
        
        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados),
            'valor_total': valor_total
        })
        
    except Exception as e:
        current_app.logger.error(f"Erro ao gerar relatório de viagens: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar relatório: {str(e)}'
        }), 500


# ========== RELATÓRIO 3: CONFERÊNCIA DE MOTORISTAS ==========

@relatorios_bp.route('/conferencia-motoristas')
@login_required
def conferencia_motoristas():
    """Exibe a tela de filtros para o relatório de conferência de motoristas"""
    
    motoristas = Motorista.query.all()
    
    return render_template(
        'relatorios/conferencia_motoristas.html',
        motoristas=motoristas
    )


@relatorios_bp.route('/conferencia-motoristas/dados', methods=['POST'])
@login_required
def dados_conferencia_motoristas():
    """Retorna os dados do relatório de conferência de motoristas em JSON"""
    
    try:
        # Receber filtros
        data_inicio = request.form.get('data_inicio')
        data_fim = request.form.get('data_fim')
        motorista_id = request.form.get('motorista_id')
        status = request.form.get('status')
        
        # Construir query
        query = Viagem.query
        
        # Aplicar filtros
        if data_inicio and data_fim:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)
            
            query = query.filter(
                db.or_(
                    db.and_(Viagem.data_inicio >= data_inicio_dt, Viagem.data_inicio <= data_fim_dt),
                    db.and_(Viagem.data_criacao >= data_inicio_dt, Viagem.data_criacao <= data_fim_dt)
                )
            )
        
        if motorista_id:
            query = query.filter_by(motorista_id=motorista_id)
        
        if status:
            query = query.filter_by(status=status)
        
        # Executar query
        viagens = query.order_by(Viagem.data_inicio.desc()).all()
        
        # Formatar dados para JSON
        dados = []
        valor_total_repasse = 0.0
        
        for viagem in viagens:
            motorista = Motorista.query.get(viagem.motorista_id) if viagem.motorista_id else None
            empresa = Empresa.query.get(viagem.empresa_id) if viagem.empresa_id else None
            planta = Planta.query.get(viagem.planta_id) if viagem.planta_id else None
            
            valor_repasse = float(viagem.valor_repasse) if viagem.valor_repasse else 0.0
            valor_total_repasse += valor_repasse
            
            dados.append({
                'id': viagem.id,
                'data_inicio': viagem.data_inicio.strftime('%d/%m/%Y %H:%M') if viagem.data_inicio else '',
                'motorista': motorista.nome if motorista else viagem.nome_motorista or 'N/A',
                'empresa': empresa.nome if empresa else 'N/A',
                'planta': planta.nome if planta else 'N/A',
                'tipo_corrida': viagem.tipo_corrida or 'N/A',
                'status': viagem.status or 'N/A',
                'placa': viagem.placa_veiculo or 'N/A',
                'qtd_passageiros': viagem.quantidade_passageiros or 0,
                'valor_repasse': valor_repasse
            })
        
        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados),
            'valor_total_repasse': valor_total_repasse
        })
        
    except Exception as e:
        current_app.logger.error(f"Erro ao gerar relatório de motoristas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar relatório: {str(e)}'
        }), 500


# ========== EXPORTAÇÃO EXCEL ==========

@relatorios_bp.route('/exportar-excel/<tipo>', methods=['POST'])
@login_required
def exportar_excel(tipo):
    """Exporta o relatório para Excel"""
    
    try:
        import json
        dados = json.loads(request.form.get('dados', '[]'))
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Definir colunas baseado no tipo
        if tipo == 'solicitacoes':
            ws.title = 'Solicitações'
            colunas = ['ID', 'Data', 'Colaborador', 'Empresa', 'Planta', 'Bloco', 
                      'Tipo Linha', 'Tipo Corrida', 'Status', 'Entrada', 'Saída', 'Valor']
            
            # Cabeçalho
            for col, valor in enumerate(colunas, 1):
                cell = ws.cell(row=1, column=col, value=valor)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            
            # Dados
            for row, item in enumerate(dados, 2):
                ws.cell(row=row, column=1, value=item.get('id'))
                ws.cell(row=row, column=2, value=item.get('data_criacao'))
                ws.cell(row=row, column=3, value=item.get('colaborador'))
                ws.cell(row=row, column=4, value=item.get('empresa'))
                ws.cell(row=row, column=5, value=item.get('planta'))
                ws.cell(row=row, column=6, value=item.get('bloco'))
                ws.cell(row=row, column=7, value=item.get('tipo_linha'))
                ws.cell(row=row, column=8, value=item.get('tipo_corrida'))
                ws.cell(row=row, column=9, value=item.get('status'))
                ws.cell(row=row, column=10, value=item.get('horario_entrada'))
                ws.cell(row=row, column=11, value=item.get('horario_saida'))
                ws.cell(row=row, column=12, value=item.get('valor'))
        
        elif tipo == 'viagens':
            ws.title = 'Viagens'
            colunas = ['ID', 'Data', 'Empresa', 'Planta', 'Motorista', 'Placa', 
                      'Tipo Corrida', 'Status', 'Passageiros', 'Valor', 'Repasse']
            
            # Cabeçalho
            for col, valor in enumerate(colunas, 1):
                cell = ws.cell(row=1, column=col, value=valor)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            
            # Dados
            for row, item in enumerate(dados, 2):
                ws.cell(row=row, column=1, value=item.get('id'))
                ws.cell(row=row, column=2, value=item.get('data_inicio'))
                ws.cell(row=row, column=3, value=item.get('empresa'))
                ws.cell(row=row, column=4, value=item.get('planta'))
                ws.cell(row=row, column=5, value=item.get('motorista'))
                ws.cell(row=row, column=6, value=item.get('placa'))
                ws.cell(row=row, column=7, value=item.get('tipo_corrida'))
                ws.cell(row=row, column=8, value=item.get('status'))
                ws.cell(row=row, column=9, value=item.get('qtd_passageiros'))
                ws.cell(row=row, column=10, value=item.get('valor'))
                ws.cell(row=row, column=11, value=item.get('valor_repasse'))
        
        elif tipo == 'motoristas':
            ws.title = 'Motoristas'
            colunas = ['ID', 'Data', 'Motorista', 'Empresa', 'Planta', 
                      'Tipo Corrida', 'Status', 'Passageiros', 'Repasse']
            
            # Cabeçalho
            for col, valor in enumerate(colunas, 1):
                cell = ws.cell(row=1, column=col, value=valor)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            
            # Dados
            for row, item in enumerate(dados, 2):
                ws.cell(row=row, column=1, value=item.get('id'))
                ws.cell(row=row, column=2, value=item.get('data_inicio'))
                ws.cell(row=row, column=3, value=item.get('motorista'))
                ws.cell(row=row, column=4, value=item.get('empresa'))
                ws.cell(row=row, column=5, value=item.get('planta'))
                ws.cell(row=row, column=6, value=item.get('tipo_corrida'))
                ws.cell(row=row, column=7, value=item.get('status'))
                ws.cell(row=row, column=8, value=item.get('qtd_passageiros'))
                ws.cell(row=row, column=9, value=item.get('valor_repasse'))
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Enviar arquivo
        filename = f'relatorio_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"Erro ao exportar Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ========== EXPORTAÇÃO PDF ==========

@relatorios_bp.route('/exportar-pdf/<tipo>', methods=['POST'])
@login_required
def exportar_pdf(tipo):
    """Exporta o relatório para PDF"""
    
    try:
        import json
        dados = json.loads(request.form.get('dados', '[]'))
        
        # Criar PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        
        # Definir dados da tabela baseado no tipo
        if tipo == 'solicitacoes':
            table_data = [['ID', 'Data', 'Colaborador', 'Empresa', 'Status', 'Valor']]
            for item in dados:
                table_data.append([
                    str(item.get('id', '')),
                    item.get('data_criacao', ''),
                    item.get('colaborador', ''),
                    item.get('empresa', ''),
                    item.get('status', ''),
                    f"R$ {item.get('valor', 0):.2f}"
                ])
        
        elif tipo == 'viagens':
            table_data = [['ID', 'Data', 'Motorista', 'Empresa', 'Status', 'Valor']]
            for item in dados:
                table_data.append([
                    str(item.get('id', '')),
                    item.get('data_inicio', ''),
                    item.get('motorista', ''),
                    item.get('empresa', ''),
                    item.get('status', ''),
                    f"R$ {item.get('valor', 0):.2f}"
                ])
        
        elif tipo == 'motoristas':
            table_data = [['ID', 'Data', 'Motorista', 'Empresa', 'Status', 'Repasse']]
            for item in dados:
                table_data.append([
                    str(item.get('id', '')),
                    item.get('data_inicio', ''),
                    item.get('motorista', ''),
                    item.get('empresa', ''),
                    item.get('status', ''),
                    f"R$ {item.get('valor_repasse', 0):.2f}"
                ])
        
        # Criar tabela
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        
        # Enviar arquivo
        filename = f'relatorio_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"Erro ao exportar PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

